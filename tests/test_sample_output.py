from pathlib import Path

from openpyxl import load_workbook


def test_sample_output_workbook_exists():
    output_path = Path("sample_outputs/fraud_analysis_output.xlsx")
    assert output_path.exists()


def test_sample_output_workbook_has_expected_sheets():
    output_path = Path("sample_outputs/fraud_analysis_output.xlsx")
    workbook = load_workbook(output_path, read_only=True)

    expected_sheets = {
        "Executive_Summary",
        "Top_Suspicious",
        "Risk_Summary",
        "Fraud_By_Risk",
        "Model_Evaluation",
        "Score_Distribution",
        "Model_Breakdown",
        "Limitations",
    }

    assert expected_sheets.issubset(set(workbook.sheetnames))